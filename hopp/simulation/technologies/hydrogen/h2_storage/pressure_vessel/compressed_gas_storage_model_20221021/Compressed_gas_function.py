# -*- coding: utf-8 -*-
"""
Created on Fri Jan 15 15:06:21 2021

@author: ppeng
"""
import openpyxl as openpyxl
import numpy as np
import math as math
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from CoolProp.CoolProp import PropsSI
from scipy.optimize import leastsq

plt.rcParams.update({'font.size': 13})

class CompressedGasFunction():

    def __init__(self, path_tankinator):
        # path to the excel spreadsheet to store material properties
        self.wb_tankinator = openpyxl.load_workbook(path_tankinator, data_only=True) #Add file name
        
        ################Other key inputs besides the main script##########################
        self.MW_H2=2.02e-03 #molecular weight of H2 in kg/mol
        
        self.Pres=int(350) #Define storage pressure in bar, Important, if you change storage pressure, make sure to change it in the corresponding tab in Tankinator and save again
        self.Temp_c = 293 #Define storage temperature in K
        self.Pin=int(30)  #Deinfe pressure out of electrolyzer in bar
        self.Tin=int(353) #Define temperature out of electrolyzer in K
        self.T_amb=int(295)
        self.Pres3=int(35) #Define outlet pressure in bar
        self.Temp3=int(353) #Define outlet temperature in K
        
        self.start_point = 10   #For setting the smallest capacity for fitting and plotting

        
        #################Economic parameters
        self.CEPCI2007 = 525.4
        self.CEPCI2001 = 397
        self.CEPCI2017 = 567.5
        
        self.CEPCI_current = 708    ####Change this value for current CEPCI 
        
        self.wage = 36
        self.maintanance = 0.03
        self.Site_preparation = 100   #Site preparation in $/kg
        
        self.Tank_manufacturing = 1.8 #self.Markup for tank manufacturing
        self.Markup = 1.5   #self.Markup for installation engineering/contingency
        
        #################Other minor input parameters############
        self.R=8.314 # gas onstant m3*Pa/(molK)
        self.Heat_Capacity_Wall = 0.92 ##wall heat capacity at 298 K in kJ/kg*K for carbon fiber composite
        self.Efficiency_comp = 0.7  #Compressor efficiency
        self.Efficiency_heater = 0.7  #Heat efficiency

    def exp_log_fit(self, var_op, capacity_1):
        
        a_op=var_op[0]
        b_op=var_op[1]
        c_op=var_op[2]

        fit_op_kg = np.exp(a_op*(np.log(capacity_1))**2-b_op*np.log(capacity_1)+c_op)

        return fit_op_kg

    def residual_op(self, var_op, capacity_1, Op_c_Costs_kg):
        
            fit_op_kg = self.exp_log_fit(var_op, capacity_1)

            return (fit_op_kg - Op_c_Costs_kg)
    
    def exp_fit(self, x, a, b):
            return a*x**b

    def calculate_max_storage_capacity(self, Wind_avai, H2_flow, Release_efficiency):
        
        H2_flow_ref = 200 #reference flow rate of steel plants in tonne/day, in case in the future it is not 200 tonne/day
        
        capacity_max = (0.8044*Wind_avai**2-57.557*Wind_avai+4483.1)*(H2_flow/H2_flow_ref)/Release_efficiency*1000  ###Total max equivalent storage capacity kg

        return capacity_max

    def calculate_max_storage_duration(self, Release_efficiency, H2_flow):

        t_discharge_hr_max = self.capacity_max/1000*Release_efficiency/H2_flow  ###This is the theoretical maximum storage duration 

        return t_discharge_hr_max
    
    #TODO keep breaking this up so we can run the model without running the curve fit
    def func(self, Wind_avai, H2_flow, cdratio, Energy_cost, cycle_number, capacity_max_spec=None, t_discharge_hr_max_spec=None):
        """
        Run the compressor and storage container cost models

        Wind_avai is only used for calculating the theoretical maximum storage capacity prior to curve fitting

        H2_flow is (I think) the rate the H2 is being removed from the tank in Tonne/day

        cdratio is the charge/discharge ratio (1 means charge rate equals the discharge rate, 2 means charge is 2x the discharge rate)

        Energy_cost is the renewable energy cost in $/kWh, or can be set to 0 to exclude energy costs

        cycle number should just be left as 1 (see compressed_all.py)
        """
        
        ##############Calculation of storage capacity from duration#############
        if 1-self.Pres3/self.Pres < 0.9:
            Release_efficiency = 1-self.Pres3/self.Pres    
        else:
            Release_efficiency = 0.9
        
        if capacity_max_spec == None:
            self.capacity_max = self.calculate_max_storage_capacity(Wind_avai, H2_flow, Release_efficiency)
        else:
            self.capacity_max = capacity_max_spec
        
        if t_discharge_hr_max_spec == None:
            self.t_discharge_hr_max = self.calculate_max_storage_duration(Release_efficiency, H2_flow)
        else:
            self.t_discharge_hr_max = t_discharge_hr_max_spec

        if self.verbose:
            print('Maximum capacity is', self.capacity_max, 'kg H2')
            print('Maximum storage duration is', self.t_discharge_hr_max, 'hr')
        
        if self.Pres > 170:
        ####Use this if use type IV tanks
            tank_type= 4
            sheet_tankinator = self.wb_tankinator['type4_rev3'] #Add Sheet name
            Vtank_c_cell = sheet_tankinator.cell(row=19, column=3)   #tank internal volume in cm3
            Vtank_c=Vtank_c_cell.value/(10**6) #tank volume in m3
            m_c_wall_cell=sheet_tankinator.cell(row=55, column=3)
            m_c_wall=m_c_wall_cell.value #Wall mass in kg
            Mtank_c=m_c_wall #TODO why is this set but not used?
            Louter_c_cell=sheet_tankinator.cell(row= 36, column= 3)
            length_outer_c=Louter_c_cell.value # outer length of tank
            Router_c_cell=sheet_tankinator.cell(row= 37, column= 3)
            radius_outer_c=Router_c_cell.value # outer radius of tank
            Cost_c_tank_cell=sheet_tankinator.cell(row=65, column=3) #Cost of one tank 
            Cost_c_tank = Cost_c_tank_cell.value   ##Cost of the tank in $/tank
            
        if self.Pres <= 170:
        ####Use this if use type I tanks
            tank_type= 1
            sheet_tankinator = self.wb_tankinator['type1_rev3'] #Add Sheet nam
            Vtank_c_cell = sheet_tankinator.cell(row=20, column=3)  ##Tank's outer volume in cm^3
            Vtank_c=Vtank_c_cell.value/(10**6) #tank volume in m3
            m_c_wall_cell=sheet_tankinator.cell(row=188, column=3)
            m_c_wall=m_c_wall_cell.value #Wall mass in kg
            Mtank_c=m_c_wall #TODO why is this set but not used?
            Louter_c_cell=sheet_tankinator.cell(row= 184, column= 3)
            length_outer_c=Louter_c_cell.value
            Router_c_cell=sheet_tankinator.cell(row= 185, column= 3)
            radius_outer_c=Router_c_cell.value
            Cost_c_tank_cell=sheet_tankinator.cell(row=193, column=3) #Cost of one tank 
            Cost_c_tank = Cost_c_tank_cell.value   ##Cost of the tank in $/tank

        self.tank_type= tank_type
        self.Vtank= Vtank_c
        self.m_H2_tank= self.Vtank*PropsSI("D", "P", self.Pres*10**5, "T", self.Temp_c, "Hydrogen")
        self.Mempty_tank= Mtank_c
        self.Router= radius_outer_c
        self.Louter= length_outer_c

        #####Define arrays for plotting and fitting  

        self.t_discharge_hr_1 = np.linspace (self.t_discharge_hr_max, self.t_discharge_hr_max/self.start_point, num=15)
        self.cost_kg = np.zeros(len(self.t_discharge_hr_1))
        cost_kg_tank = np.zeros(len(self.t_discharge_hr_1))
        cost_kg_comp = np.zeros(len(self.t_discharge_hr_1))
        cost_kg_ref = np.zeros(len(self.t_discharge_hr_1))
        cost_kg_heat = np.zeros(len(self.t_discharge_hr_1))
        self.number_of_tanks = np.zeros(len(self.t_discharge_hr_1))
        self.capacity_1 = np.zeros(len(self.t_discharge_hr_1))
        self.Op_c_Costs_kg = np.zeros(len(self.t_discharge_hr_1))
        self.total_energy_used_kwh = np.zeros(len(self.t_discharge_hr_1))
        
        ###################################################################################################
        ###################################################################################################
        ###################################################################################################
        ###############Starting detailed calculations#################################
        ###############Stage 1 calculations#################################
        
        for i in range (0,len(self.t_discharge_hr_1-1)):
            t_discharge_hr = self.t_discharge_hr_1 [i]
            capacity=H2_flow*t_discharge_hr*1000/Release_efficiency #Maximum capacity in kg H2
            
            self.capacity_1 [i] = capacity
            
            rgas=PropsSI("D", "P", self.Pres*10**5, "T", self.Temp_c, "Hydrogen") #h2 density in kg/m3 under storage conditions
            H2_c_mass_gas_tank = Vtank_c*rgas  #hydrogen mass per tank in kg
            H2_c_mass_tank = H2_c_mass_gas_tank  #Estimation of H2 amount per tank in kg
            self.single_tank_h2_capacity_kg = H2_c_mass_tank
            
            number_c_of_tanks = np.ceil(capacity/H2_c_mass_tank)
            self.number_of_tanks[i]= number_c_of_tanks
            H2_c_Cap_Storage= H2_c_mass_tank*(number_c_of_tanks-1)+capacity%H2_c_mass_tank  ####This will be useful when changing to assume all tanks are full, but will cause the model to not perform well for small scales, where 1 tank makes a large difference
            
            #################Energy balance for adsorption (state 1 to state 2)########
            self.t_charge_hr=t_discharge_hr * (1/cdratio)
            t_precondition_hr=self.t_charge_hr  #correcting first cycle, useful to size based on maximum power and also when calculating the operational cost
            m_c_flow_rate_1_2 = H2_c_Cap_Storage/t_precondition_hr/3600 #mass flow rate in kg/s
            Temp2=self.Temp_c
            Temp1_gas=self.Tin
            Temp1_solid=self.T_amb
            Pres2=self.Pres*10**5
            Pres1=self.Pin*10**5
            H_c_1_spec_g=PropsSI("H", "P", Pres1, "T", Temp1_gas, "Hydrogen")/1000 #specific enthalpy of the gas under T1 P1 in kJ/kg
            H_c_2_spec_g=PropsSI("H", "P", Pres2, "T", Temp2, "Hydrogen")/1000 #specific enthalpy of the gas under T2 P2 in kJ/kg
            H_c_1_gas = H2_c_Cap_Storage*H_c_1_spec_g
            H_c_2_gas = H2_c_Cap_Storage*H_c_2_spec_g
            deltaE_c_H2_1_2 = H_c_2_gas-H_c_1_gas
            deltaE_c_Uwall_1_2 = self.Heat_Capacity_Wall*(Temp2-Temp1_solid)*m_c_wall*number_c_of_tanks #Net energy/enthalpy change of adsorbent in kJ
            deltaE_c_net_1_2 = deltaE_c_H2_1_2 + deltaE_c_Uwall_1_2 #Net energy/enthalpy change in kJ
            deltaP_c_net_1_2 = deltaE_c_net_1_2/self.t_charge_hr/3600  #Net power change in kW
                
            #################Energy balance for desorption (state 2 to state 3)########
            Temp3_gas=self.Temp3
            Temp3_solid = Temp2
            self.Pres3=self.Pres3
            Pres3_tank=self.Pres*(1-Release_efficiency)*10**5*10
            H_c_3_spec_g_fuel_cell=PropsSI("H", "P", self.Pres3, "T", Temp3_gas, "Hydrogen")/1000 #specific enthalpy of the released gas in kJ/kg
            H_c_3_spec_g_tank=PropsSI("H", "P", Pres3_tank, "T", Temp2, "Hydrogen")/1000 #specific enthalpy of the remaining free volume gas in kJ/kg
            H_c_3_gas = H2_c_Cap_Storage*Release_efficiency*H_c_3_spec_g_fuel_cell+H2_c_Cap_Storage*(1-Release_efficiency)*H_c_3_spec_g_tank  #Total gas phase enthalpy in stage 3 in kJ
            deltaE_c_H2_2_3=H_c_3_gas-H_c_2_gas #Total h2 enthalpy change in kJ
            deltaE_c_Uwall_2_3 = self.Heat_Capacity_Wall*(Temp3_solid-Temp2)*m_c_wall*number_c_of_tanks  #kJ
            deltaE_c_net_2_3 = deltaE_c_H2_2_3+deltaE_c_Uwall_2_3 # Net enthalpy change during desorption
            detlaP_c_net_2_3 = deltaE_c_net_2_3/t_discharge_hr/3600
            
            ###############Energy balance for adsorption (state 4 to state 2)##########
            m_c_flow_rate_4_2 = H2_c_Cap_Storage*Release_efficiency/self.t_charge_hr/3600
            Temp4_tank=Temp2
            Temp4_gas = self.Tin #TODO why set but not used?
            Pres4=self.Pres3    #TODO why set but not used?
            Pres4_tank = Pres3_tank
            H_c_4_spec_g_electrolyzer=PropsSI("H", "P", self.Pin, "T", self.Tin, "Hydrogen")/1000 #specific enthalpy of the released gas in kJ/kg
            H_c_4_spec_g_tank=PropsSI("H", "P", Pres4_tank, "T", Temp2-5, "Hydrogen")/1000 #specific enthalpy of the remaining free volume gas in kJ/kg
            H_c_4_gas = H2_c_Cap_Storage*Release_efficiency*H_c_4_spec_g_electrolyzer+H2_c_Cap_Storage*(1-Release_efficiency)*H_c_4_spec_g_tank  #Total gas phase enthalpy in stage 3 in kJ
            deltaE_c_H2_4_2=H_c_2_gas-H_c_4_gas #Total h2 enthalpy change in kJ
            deltaE_c_Uwall_4_2 = self.Heat_Capacity_Wall*(Temp2-Temp4_tank)*m_c_wall*number_c_of_tanks  #kJ
            deltaE_c_net_4_2 = deltaE_c_H2_4_2 +deltaE_c_Uwall_4_2 # Net enthalpy change during desorption
            deltaP_c_net_4_2 = deltaE_c_net_4_2/self.t_charge_hr/3600
            
            
            ########################################Costs for cycle 1 adsorption##################################
            
            ########################################CAPITAL COSTS (sized based on cycle 1 requirements)###########################################
            
            ###############################Compressor costs ### axial/centrifugal
            if self.Pres>=self.Pin:
                K=PropsSI("ISENTROPIC_EXPANSION_COEFFICIENT", "P", self.Pin*10**5, "T", self.Tin, "Hydrogen")
                P2nd = self.Pin*(self.Pres/self.Pin)**(1/3)
                P3rd = self.Pin*(self.Pres/self.Pin)**(1/3)*(self.Pres/self.Pin)**(1/3)
                work_c_comp_1 = K/(K-1)*self.R*self.Tin/self.MW_H2*((P2nd/self.Pin)**((K-1)/K)-1)
                work_c_comp_2 = K/(K-1)*self.R*self.Tin/self.MW_H2*((P3rd/P2nd)**((K-1)/K)-1)
                work_c_comp_3 = K/(K-1)*self.R*self.Tin/self.MW_H2*((self.Pres/P3rd)**((K-1)/K)-1)
                Work_c_comp = work_c_comp_1+work_c_comp_2+work_c_comp_3
                # Work_c_comp=K/(K-1)*self.R*self.Tin/self.MW_H2*((self.Pres/self.Pin)**((K-1)/K)-1) #mechanical energy required for compressor in J/kg (single stage)
                Power_c_comp_1_2=Work_c_comp/1000*m_c_flow_rate_1_2 #mechanical power of the pump in kW
                Power_c_comp_4_2=Work_c_comp/1000*m_c_flow_rate_4_2
                A_c_comp_1_2 = Power_c_comp_1_2/self.Efficiency_comp  #total power in kW
                A_c_comp_4_2 = Power_c_comp_4_2/self.Efficiency_comp #total power in kW
                if A_c_comp_1_2>=A_c_comp_4_2:
                    A_c_comp=A_c_comp_1_2
                else:
                    A_c_comp=A_c_comp_4_2
                    
                # print ('work of compressor is', Work_c_comp,'J/kg')
                # print ('Adjusted storage capacity is', H2_c_Cap_Storage, 'kg')
                # print ('flow rate is', m_c_flow_rate_1_2, 'and', m_c_flow_rate_4_2, 'kg/s')
                # print('Total fluid power of compressor', A_c_comp, 'kW')
                Number_c_Compressors=np.floor(A_c_comp/3000) #Number of compressors excluding the last one
                A_c_comp_1 = A_c_comp%3000  #power of the last compressor
                # print('Number of compressors', Number_c_Compressors+1)
                k1=2.2897
                k2=1.3604
                k3=-0.1027
                Compr_c_Cap_Cost=(10**(k1+k2*np.log10(3000)+k3*(np.log10(3000))**2))*Number_c_Compressors
                Compr_c_Cap_Cost_1=(10**(k1+k2*np.log10(A_c_comp_1)+k3*(np.log10(A_c_comp_1))**2))

                compressor_energy_used_1 = Work_c_comp*H2_c_Cap_Storage*2.8e-7
                compressor_energy_used_2 = Work_c_comp*H2_c_Cap_Storage*Release_efficiency*2.8e-7

                Compr_c_Energy_Costs_1 = compressor_energy_used_1*Energy_cost  #compressor electricity cost in cycle 1
                Compr_c_Energy_Costs_2 = compressor_energy_used_2*Energy_cost #compressor electricity cost assuming in regular charging cycle 
                
                Total_c_Compr_Cap_Cost = Compr_c_Cap_Cost + Compr_c_Cap_Cost_1
                Total_c_Compr_Cap_Cost = Total_c_Compr_Cap_Cost*(self.CEPCI_current/self.CEPCI2001)  ##Inflation
            else:
                Power_c_comp_1_2=0 #mechanical power of the pump in kW
                Power_c_comp_4_2=0
                A_c_comp_1_2 = 0  #total power in kW
                A_c_comp_4_2 = 0 #total power in kW
                Work_c_comp = 0
                Compr_c_Cap_Cost = 0
                compressor_energy_used_1 = 0
                compressor_energy_used_2 = 0
                Compr_c_Energy_Costs_1 = 0
                Compr_c_Energy_Costs_2 = 0
                Total_c_Compr_Cap_Cost = 0

            self.total_compressor_energy_used_kwh = compressor_energy_used_1 #+ compressor_energy_used_2

            # print ('Compressor energy cost is $', Compr_c_Energy_Costs)
            # print ('refrigeration capcost for compressor is $')
            # print('compressor capcost is $', Total_c_Compr_Cap_Cost)
            # print("----------")
            
            ########################################Costs associated with storage tanks
            
            # print("Number of tanks is: ", number_c_of_tanks)
            Storage_c_Tank_Cap_Costs=Cost_c_tank*number_c_of_tanks*self.Tank_manufacturing
            Storage_c_Tank_Cap_Costs = Storage_c_Tank_Cap_Costs*(self.CEPCI_current/self.CEPCI2007)  ##Inflation
            # print('Capcost for storage tank is', Storage_c_Tank_Cap_Costs)
            # print("----------")
            
            ###############################Refrigeration costs estimation adsorption process     
            # print ('pre-conditioning time is', round (t_precondition_hr), 'hr')
            H2_Cap=(capacity)    
            Ref_c_P_net_1_2=-(deltaP_c_net_1_2-Power_c_comp_1_2)  #Refrigeration power in kW from state 1 to state 2 (precondition)
            Ref_c_P_net_4_2=-(deltaP_c_net_4_2-Power_c_comp_4_2)  #Refrigeration power in kW from state 1 to state 2 (normal charging)
            if Ref_c_P_net_1_2 >= Ref_c_P_net_4_2:    
                Net_c_Cooling_Power_Adsorption=Ref_c_P_net_1_2 #Net refrigeration power in kW
            else:
                Net_c_Cooling_Power_Adsorption=Ref_c_P_net_4_2
            # print ('Net Cooling power for refrigeration sizing is',Net_c_Cooling_Power_Adsorption,'kW') #Cooling power in kW
            
            if Net_c_Cooling_Power_Adsorption < 1000:
                A1=-3.53E-09
                A2=-9.94E-06
                A3=3.30E-03
                nc=(A1*(self.Temp_c**3))+(A2*(self.Temp_c**2))+A3*self.Temp_c #Carnot efficiency factor
                COP=(self.Temp_c/(318-self.Temp_c))*nc #Coefficient of performance
                B1=24000
                B2=3500
                B3=0.9
                Total_c_Refrig_Cap_Costs_adsorption=(B1+(B2*(Net_c_Cooling_Power_Adsorption/COP)**B3))*(self.CEPCI_current/550.8)   
            else:
                Total_c_Refrig_Cap_Costs_adsorption=2*10**11*self.Temp_c**-2.077*(Net_c_Cooling_Power_Adsorption/1000)**0.6
                Total_c_Refrig_Cap_Costs_adsorption = Total_c_Refrig_Cap_Costs_adsorption*(self.CEPCI_current/self.CEPCI2017) 
            
            ####Utility for refrigeration
            Utility_c_ref = 4.07*10**7*self.Temp_c**(-2.669) #Utility in $/GJ, here, the utility is mostly for energy assumes 16.8 $/GJ (57 $/MWh)
            # Utility_c_refrigeration_1 = (self.CEPCI_current/self.CEPCI2017)*Utility_c_ref*-(deltaE_c_net_1_2-Work_c_comp*H2_c_Cap_Storage/1000)/1e6  
            energy_consumption_refrigeration_1_kj = -(deltaE_c_net_1_2-Work_c_comp*H2_c_Cap_Storage/1000) # in kJ
            
            Utility_c_refrigeration_1 = (Energy_cost/0.057)*Utility_c_ref*energy_consumption_refrigeration_1_kj/1e6  # changed based on discussion with original author 20221216, energy separated out 20230317
            # print ('refrigerator capital cost for adsorption is $', Total_c_Refrig_Cap_Costs_adsorption)    
            # print("------------")
            
            # Utility_c_refrigeration_2 = (self.CEPCI_current/self.CEPCI2017)*Utility_c_ref*-(deltaE_c_net_4_2-Work_c_comp*H2_c_Cap_Storage*Release_efficiency/1000)/1e6 
            energy_consumption_refrigeration_2_kj = -(deltaE_c_net_4_2-Work_c_comp*H2_c_Cap_Storage*Release_efficiency/1000) # in kJ
            Utility_c_refrigeration_2 = (Energy_cost/0.057)*Utility_c_ref*energy_consumption_refrigeration_2_kj/1e6  # changed based on discussion with original author 20221216, energy separated out 20230317
            
            # specify energy usage separately so energy usage can be used externally if desired
            joule2watthour = 1.0/3600.0 # 3600 joules in a watt hour (as also 3600 kJ in a kWh)
            energy_consumption_refrigeration_1_kwh = energy_consumption_refrigeration_1_kj*joule2watthour
            energy_consumption_refrigeration_2_kwh = energy_consumption_refrigeration_2_kj*joule2watthour
            self.total_refrigeration_energy_used_kwh = energy_consumption_refrigeration_1_kwh #+ energy_consumption_refrigeration_2_kwh
            
            if self.total_refrigeration_energy_used_kwh < 0:
                raise(ValueError("energy usage must be greater than 0"))
            ###############################Heating costs desorption process   
            k1=6.9617
            k2=-1.48
            k3=0.3161
            Net_c_Heating_Power_Desorption=detlaP_c_net_2_3/self.Efficiency_heater ## steam boiler power at 0.7 efficiency in kW
            Number_c_Heaters=np.floor(Net_c_Heating_Power_Desorption/9400) #Number of compressors excluding the last one
            Heater_c_Power_1 = Net_c_Heating_Power_Desorption%9400  #power of the last compressor
            # print('Number of heaters', Number_c_Heaters+1)
            Heater_c_Cap_Cost=(10**(k1+k2*np.log10(9400)+k3*(np.log10(9400))**2))*Number_c_Heaters
            if Heater_c_Power_1 < 1000:
                Heater_c_Cap_Cost_1=(10**(k1+k2*np.log10(1000)+k3*(np.log10(1000))**2))*(Heater_c_Power_1/1000)
            else:
                Heater_c_Cap_Cost_1=(10**(k1+k2*np.log10(Heater_c_Power_1)+k3*(np.log10(Heater_c_Power_1))**2))
            Total_c_Heater_Cap_Cost = Heater_c_Cap_Cost + Heater_c_Cap_Cost_1
            Total_c_Heater_Cap_Cost = Total_c_Heater_Cap_Cost *(self.CEPCI_current/self.CEPCI2001)  ##Inflation #TODO make inflation optional per user input
            
            Utility_c_Heater = 0 # Jared Thomas set to zero as per discussion with Peng Peng through Abhineet Gupta 20221215 was 13.28*deltaE_c_net_2_3/1e6 #$13.28/GJ for low pressure steam
            self.total_heating_energy_used_kwh = Net_c_Heating_Power_Desorption*t_discharge_hr
            Total_c_Heating_Energy_Costs = self.total_heating_energy_used_kwh*Energy_cost
            
            # print('heater capcost is $', Total_c_Heater_Cap_Cost)
            
            ########################################Operational costs (sized based on cycle 1 requirements)###########################################
            Op_c_Costs_1 = Compr_c_Energy_Costs_1 + Utility_c_refrigeration_1 + Utility_c_Heater + Total_c_Heating_Energy_Costs
            Op_c_Costs_2 = Compr_c_Energy_Costs_2 + Utility_c_refrigeration_2 + Utility_c_Heater + Total_c_Heating_Energy_Costs
            Total_c_Cap_Costs = Storage_c_Tank_Cap_Costs + Total_c_Refrig_Cap_Costs_adsorption + Total_c_Compr_Cap_Cost + Total_c_Heater_Cap_Cost
            
            # Op_c_Costs = (Op_c_Costs_1 + Op_c_Costs_2 * (cycle_number-1)+self.maintanance*Total_c_Cap_Costs+self.wage*360*2)/cycle_number/capacity
            #TODO check this. I changed the 2 to a 24 because it looks like it should be working hours in a year.
            Op_c_Costs = ((Op_c_Costs_1 + Op_c_Costs_2*(cycle_number-1) + self.maintanance*Total_c_Cap_Costs + self.wage*360*2)/cycle_number) # checked, this was divided by capacity, but Peng Peng confirmed it was duplicating the following divisions by capacity
            
            ######################writing costs#####################################################
            self.cost_kg[i] = (Total_c_Cap_Costs/capacity + self.Site_preparation)*self.Markup
            cost_kg_tank[i] = Storage_c_Tank_Cap_Costs/capacity
            cost_kg_comp[i] = Total_c_Compr_Cap_Cost/capacity
            cost_kg_ref[i] = Total_c_Refrig_Cap_Costs_adsorption/capacity    
            cost_kg_heat[i] = Total_c_Heater_Cap_Cost/capacity
            self.Op_c_Costs_kg[i] = Op_c_Costs/capacity
            # print("\n Pressure Vessel Costs: ")
            # print("cost_kg ")
            # print("cost_kg_tank ")
            # print("cost_kg_comp ")
            # print("cost_kg_ref ")
            # print("cost_kg_heat ")
            ######################################## Total Energy Use (kWh) ######################
            self.total_energy_used_kwh[i] = self.total_compressor_energy_used_kwh + self.total_heating_energy_used_kwh + self.total_refrigeration_energy_used_kwh
        
        self.curve_fit()

    def curve_fit(self):

        ################### plot prep ###########
        self.plot_range = range(int(np.min(self.capacity_1)), int(np.max(self.capacity_1)), 100)

        ###################Fitting capital####################################################  

        var_cap = [0.01 ,0.5, 5]   #Initial guesses for the parameters, can be flexible
        
        varfinal_cap_fitted, success = leastsq(self.residual_op, var_cap, args=(self.capacity_1, self.cost_kg), maxfev=100000)
        
        self.a_cap_fit=varfinal_cap_fitted[0]
        self.b_cap_fit=varfinal_cap_fitted[1]
        self.c_cap_fit=varfinal_cap_fitted[2] 

        if self.verbose:
            print ('a_cap is', self.a_cap_fit)
            print ('b_cap is', self.b_cap_fit)
            print ('c_cap is', self.c_cap_fit)
            print ('***********')

        self.fitted_capex = self.exp_log_fit(varfinal_cap_fitted, self.plot_range)      
        
        # popt, pcov = curve_fit(self.exp_fit, self.capacity_1, self.cost_kg, maxfev=100000)
        
        # self.a_cap_fit=popt[0]
        # self.b_cap_fit=popt[1]
        
        # print ('a is', self.a_cap_fit)
        # print ('b is', self.b_cap_fit)
        # print ('***********')
        
        # self.fitted_kg = self.exp_fit(self.plot_range,self.a_cap_fit,self.b_cap_fit)
        
        ####################### fitting OpEx #################################
        var_op= [0.01 ,0.5, 5]   #Initial guesses for the parameters, can be flexible
        
        varfinal_op_fitted,success = leastsq(self.residual_op, var_op, args=(self.capacity_1, self.Op_c_Costs_kg), maxfev=100000)
        
        self.a_op_fit=varfinal_op_fitted[0]
        self.b_op_fit=varfinal_op_fitted[1]
        self.c_op_fit=varfinal_op_fitted[2] 

        if self.verbose:
            print ('a_op is', self.a_op_fit)
            print ('b_op is', self.b_op_fit)
            print ('c_op is', self.c_op_fit)
            print ('***********')

        self.fitted_op_kg = self.exp_log_fit(varfinal_op_fitted, self.plot_range)
  
        ##################### Fit energy usage ################################
        self.energy_coefficients = np.polyfit(self.capacity_1, self.total_energy_used_kwh, 1)
        self.energy_function = np.poly1d(self.energy_coefficients) # kWh
        self.fit_energy_wrt_capacity_kwh = self.energy_function(self.plot_range)
    
    def plot(self):

        fig, ax = plt.subplots(2,2, sharex=True, figsize=(10,6))
        
        ##################### CAPEX #######################
        ax[0,0].scatter(self.capacity_1*1E-3, self.cost_kg, color='r', label = 'Calc')
        ax[0,0].plot(np.asarray(self.plot_range)*1E-3, self.fitted_capex, label = 'Fit')
        # ax[0,0].plot(self.capacity_1,cost_kg_tank, color='b', label = 'tank')
        # ax[0,0].plot(self.capacity_1,cost_kg_comp, color='c', label = 'compressor')
        # ax[0,0].plot(self.capacity_1,cost_kg_ref, color='m', label = 'refrigeration')
        # ax[0,0].plot(self.capacity_1,cost_kg_heat, color='y', label = 'heater')   
        
        a_disp=np.round(self.a_cap_fit, 2)
        b_disp=np.round(self.b_cap_fit, 2)
        # plt.ylim(0,np.amax(self.cost_kg)*2)
        # equation_cap = 'y='+str(a_disp)+'x'+'^'+str(b_disp)
        a_cap_fit_disp=np.round(self.a_cap_fit, 2)
        b_cap_fit_disp=np.round(self.b_cap_fit, 2)
        c_cap_fit_disp=np.round(self.c_cap_fit, 2) 
        equation_cap = 'y='+'exp('+str(a_cap_fit_disp)+'(ln(x))^2\n-'+str(b_cap_fit_disp)+'ln(x)+' + str(c_cap_fit_disp) +')'

        ax[0,0].annotate(equation_cap, xy=(np.amax(self.capacity_1)*1E-3*0.4 , np.amax(self.cost_kg)*0.8)) 
        
        ax[0,0].set_ylabel("CAPEX ($/kg)")
        ax[0,0].legend(loc='best', frameon=False)
        # plt.legend(loc='best')
        # ax[0,0].title('Capital')

        ##################### OPEX ############################
        a_op_fit_disp=np.round(self.a_op_fit, 2)
        b_op_fit_disp=np.round(self.b_op_fit, 2)
        c_op_fit_disp=np.round(self.c_op_fit, 2) 
        
        equation_op = 'y='+'exp('+str(a_op_fit_disp)+'(ln(x))^2\n-'+str(b_op_fit_disp)+'ln(x)+' + str(c_op_fit_disp) +')'
        
        ax[0,1].plot(np.asarray(self.plot_range)*1E-3, self.fitted_op_kg, label = 'Fit')
        ax[0,1].scatter(self.capacity_1*1E-3, self.Op_c_Costs_kg, color='r', label = 'Calc')     
        ax[0,1].set_ylabel("OPEX ($/kg)")
        ax[0,1].annotate(equation_op, xy=(np.amax(self.capacity_1)*1E-3*0.2 , np.amax(self.Op_c_Costs_kg)*0.4))     
        ax[0,1].legend(loc='best', frameon=False)
        # plt.legend(loc='best')
        # ax[0,1].title('Annual operational')

        ################## Energy ######################  
        ax[1,1].plot(np.asarray(self.plot_range)*1E-3, self.fit_energy_wrt_capacity_kwh*1E-6, label = 'Fit')
        ax[1,1].scatter(self.capacity_1*1E-3, self.total_energy_used_kwh*1E-6, color='r', label = 'Calc')    
        ax[1,1].set_xlabel("Capacity (Tonnes H2)")    
        ax[1,1].set_ylabel("Energy Use (GWh)")
        
        equation_energy = 'y='+str(round(self.energy_coefficients[0],2)) +'x+'+str(round(self.energy_coefficients[1],2)) 
        ax[1,1].annotate(equation_energy, xy=(3000, 5))     
        ax[1,1].legend(loc='best', frameon=False)
        ax[1,1].legend(loc='best', frameon=False)
        # ax[1,1].title('Annual operational')

        ################ Wrap Up ######################
        
        ax[1,0].set_xlabel("Capacity (Tonnes H2)")    

        plt.tight_layout()
        plt.show()
