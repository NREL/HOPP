# -*- coding: utf-8 -*-
"""
Created on Fri Jan 15 15:06:21 2021

@author: ppeng
"""
import openpyxl as openpyxl
from openpyxl import Workbook
import numpy as np
import pandas as pd
import math as math
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from CoolProp.CoolProp import PhaseSI, PropsSI, get_global_param_string
from scipy.optimize import leastsq
plt.rcParams.update({'font.size': 13})

def func (Wind_avai, H2_flow, cdratio, Energy_cost, cycle_number, path_tankinator):

    ##########Compressed gas storage
    ####################################################
    ####################################################
    ####################################################
    ####################################################
    ##################################Reading excel######################
    ############Importing tankinator##########################
    
    # path to the excel spreadsheet to store material properties
    wb_tankinator = openpyxl.load_workbook(path_tankinator, data_only=True) #Add file name
    
    
    ################Other key inputs besides the main script##########################
    MW_H2=2.02e-03 #molecular weight of H2 in kg/mol
    
    Pres=int(350) #Define storage pressure in bar, Important, if you change storage pressure, make sure to change it in the corresponding tab in Tankinator and save again
    Temp_c = 293 #Define storage temperature in K
    Pin=int(30)  #Deinfe pressure out of electrolyzer in bar
    Tin=int(353) #Define temperature out of electrolyzer in K
    T_amb=int(295)
    Pres3=int(35) #Define outlet pressure in bar
    Temp3=int(353) #Define outlet temperature in K
    
    
    start_point = 10   #For setting the smallest capacity for fitting and plotting
    ##############Calculation of storage capacity from duration#############
    if 1-Pres3/Pres < 0.9:
        Release_efficiency = 1-Pres3/Pres    
    else:
        Release_efficiency = 0.9
    
    H2_flow_ref = 200 #reference flow rate of steel plants in tonne/day, in case in the future it is not 200 tonne/day
    
    global capacity_max
    capacity_max = (0.8044*Wind_avai**2-57.557*Wind_avai+4483.1)*(H2_flow/H2_flow_ref)/Release_efficiency*1000  ###Total max equivalent storage capacity kg
    
    global t_discharge_hr_max
    t_discharge_hr_max = capacity_max/1000*Release_efficiency/H2_flow  ###This is the theoretical maximum storage duration        
    
    print('Maximum capacity is', capacity_max, 'kg H2')
    print('Maximum storage duration is', t_discharge_hr_max, 'hr')
    
    #################Economic parameters
    CEPCI2007 = 525.4
    CEPCI2001 = 397
    CEPCI2017 = 567.5
    
    CEPCI_current = 708    ####Change this value for current CEPCI 
    
    wage = 36
    maintanance = 0.03
    Site_preparation = 100   #Site preparation in $/kg
    
    Tank_manufacturing = 1.8 #Markup for tank manufacturing
    Markup = 1.5   #Markup for installation engineering/contingency
    
    #################Other minor input parameters############
    R=8.314 # gas onstant m3*Pa/(molK)
    Heat_Capacity_Wall = 0.92 ##wall heat capacity at 298 K in kJ/kg*K for carbon fiber composite
    Efficiency_comp = 0.7  #Compressor efficiency
    Efficiency_heater = 0.7  #Heat efficiency
    
    
    if Pres > 170:
    ####Use this if use type IV tanks
        sheet_tankinator = wb_tankinator['type4_rev3'] #Add Sheet name
        Vtank_c_cell = sheet_tankinator.cell(row=19, column=3)   #tank internal volume in cm3
        Vtank_c=Vtank_c_cell.value/(10**6) #tank volume in m3
        m_c_wall_cell=sheet_tankinator.cell(row=55, column=3)
        m_c_wall=m_c_wall_cell.value #Wall mass in kg
        Mtank_c=m_c_wall
        Cost_c_tank_cell=sheet_tankinator.cell(row=65, column=3) #Cost of one tank 
        Cost_c_tank = Cost_c_tank_cell.value   ##Cost of the tank in $/tank
        
    
    if Pres <= 170:
    ####Use this if use type I tanks
        sheet_tankinator = wb_tankinator['type1_rev3'] #Add Sheet nam
        Vtank_c_cell = sheet_tankinator.cell(row=20, column=3)  ##Tank's outer volume in cm^3
        Vtank_c=Vtank_c_cell.value/(10**6) #tank volume in m3
        m_c_wall_cell=sheet_tankinator.cell(row=188, column=3)
        m_c_wall=m_c_wall_cell.value #Wall mass in kg
        Mtank_c=m_c_wall
        Cost_c_tank_cell=sheet_tankinator.cell(row=193, column=3) #Cost of one tank 
        Cost_c_tank = Cost_c_tank_cell.value   ##Cost of the tank in $/tank
            
    #####Define arrays for plotting and fitting

    global t_discharge_hr_1
    global cost_kg
    global capacity_1    

    t_discharge_hr_1 = np.linspace (t_discharge_hr_max, t_discharge_hr_max/start_point, num=15)
    cost_kg = np.zeros(len(t_discharge_hr_1))
    cost_kg_tank = np.zeros(len(t_discharge_hr_1))
    cost_kg_comp = np.zeros(len(t_discharge_hr_1))
    cost_kg_ref = np.zeros(len(t_discharge_hr_1))
    cost_kg_heat = np.zeros(len(t_discharge_hr_1))
    capacity_1 = np.zeros(len(t_discharge_hr_1))
    Op_c_Costs_kg = np.zeros(len(t_discharge_hr_1))
    
    ###################################################################################################
    ###################################################################################################
    ###################################################################################################
    ###############Starting detailed calculations#################################
    ###############Stage 1 calculations#################################
    
    for i in range (0,len(t_discharge_hr_1-1)):
        t_discharge_hr = t_discharge_hr_1 [i]
        capacity=H2_flow*t_discharge_hr*1000/Release_efficiency #Maximum capacity in kg H2
        capacity_1 [i] = capacity
        
        rgas=PropsSI("D", "P", Pres*10**5, "T", Temp_c, "Hydrogen") #h2 density in kg/m3 under storage conditions
        H2_c_mass_gas_tank = Vtank_c*rgas  #hydrogen mass per tank in kg
        H2_c_mass_tank = H2_c_mass_gas_tank  #Estimation of H2 amount per tank in kg
        # print("total hydrogen stored per tank is", H2_c_mass_tank, "kg")

        
        number_c_of_tanks = np.ceil(capacity/H2_c_mass_tank)
        H2_c_Cap_Storage= H2_c_mass_tank*(number_c_of_tanks-1)+capacity%H2_c_mass_tank  ####This will be useful when changing to assume all tanks are full, but will cause the model to not perform well for small scales, where 1 tank makes a large difference

        
        #################Energy balance for adsorption (state 1 to state 2)########
        global t_charge_hr
        t_charge_hr=t_discharge_hr * (1/cdratio)
        t_precondition_hr=t_charge_hr  #correcting first cycle, useful to size based on maximum power and also when calculating the operational cost
        m_c_flow_rate_1_2 = H2_c_Cap_Storage/t_precondition_hr/3600 #mass flow rate in kg/s
        Temp2=Temp_c
        Temp1_gas=Tin
        Temp1_solid=T_amb
        Pres2=Pres*10**5
        Pres1=Pin*10**5
        H_c_1_spec_g=PropsSI("H", "P", Pres1, "T", Temp1_gas, "Hydrogen")/1000 #specific enthalpy of the gas under T1 P1 in kJ/kg
        H_c_2_spec_g=PropsSI("H", "P", Pres2, "T", Temp2, "Hydrogen")/1000 #specific enthalpy of the gas under T2 P2 in kJ/kg
        H_c_1_gas = H2_c_Cap_Storage*H_c_1_spec_g
        H_c_2_gas = H2_c_Cap_Storage*H_c_2_spec_g
        deltaE_c_H2_1_2 = H_c_2_gas-H_c_1_gas
        deltaE_c_Uwall_1_2 = Heat_Capacity_Wall*(Temp2-Temp1_solid)*m_c_wall*number_c_of_tanks #Net energy/enthalpy change of adsorbent in kJ
        deltaE_c_net_1_2 = deltaE_c_H2_1_2 + deltaE_c_Uwall_1_2 #Net energy/enthalpy change in kJ
        deltaP_c_net_1_2 = deltaE_c_net_1_2/t_charge_hr/3600  #Net power change in kW
            
            
        #################Energy balance for desorption (state 2 to state 3)########
        Temp3_gas=Temp3
        Temp3_solid = Temp2
        Pres3=Pres3
        Pres3_tank=Pres*(1-Release_efficiency)*10**5*10
        H_c_3_spec_g_fuel_cell=PropsSI("H", "P", Pres3, "T", Temp3_gas, "Hydrogen")/1000 #specific enthalpy of the released gas in kJ/kg
        H_c_3_spec_g_tank=PropsSI("H", "P", Pres3_tank, "T", Temp2, "Hydrogen")/1000 #specific enthalpy of the remaining free volume gas in kJ/kg
        H_c_3_gas = H2_c_Cap_Storage*Release_efficiency*H_c_3_spec_g_fuel_cell+H2_c_Cap_Storage*(1-Release_efficiency)*H_c_3_spec_g_tank  #Total gas phase enthalpy in stage 3 in kJ
        deltaE_c_H2_2_3=H_c_3_gas-H_c_2_gas #Total h2 enthalpy change in kJ
        deltaE_c_Uwall_2_3 = Heat_Capacity_Wall*(Temp3_solid-Temp2)*m_c_wall*number_c_of_tanks  #kJ
        deltaE_c_net_2_3 = deltaE_c_H2_2_3+deltaE_c_Uwall_2_3 # Net enthalpy change during desorption
        detlaP_c_net_2_3 = deltaE_c_net_2_3/t_discharge_hr/3600
        
        ###############Energy balance for adsorption (state 4 to state 2)##########
        m_c_flow_rate_4_2 = H2_c_Cap_Storage*Release_efficiency/t_charge_hr/3600
        Temp4_tank=Temp2
        Temp4_gas = Tin
        Pres4=Pres3
        Pres4_tank = Pres3_tank
        H_c_4_spec_g_electrolyzer=PropsSI("H", "P", Pin, "T", Tin, "Hydrogen")/1000 #specific enthalpy of the released gas in kJ/kg
        H_c_4_spec_g_tank=PropsSI("H", "P", Pres4_tank, "T", Temp2-5, "Hydrogen")/1000 #specific enthalpy of the remaining free volume gas in kJ/kg
        H_c_4_gas = H2_c_Cap_Storage*Release_efficiency*H_c_4_spec_g_electrolyzer+H2_c_Cap_Storage*(1-Release_efficiency)*H_c_4_spec_g_tank  #Total gas phase enthalpy in stage 3 in kJ
        deltaE_c_H2_4_2=H_c_2_gas-H_c_4_gas #Total h2 enthalpy change in kJ
        deltaE_c_Uwall_4_2 = Heat_Capacity_Wall*(Temp2-Temp4_tank)*m_c_wall*number_c_of_tanks  #kJ
        deltaE_c_net_4_2 = deltaE_c_H2_4_2 +deltaE_c_Uwall_4_2 # Net enthalpy change during desorption
        deltaP_c_net_4_2 = deltaE_c_net_4_2/t_charge_hr/3600
        
        
        ########################################Costs for cycle 1 adsorption##################################
        
        ########################################CAPITAL COSTS (sized based on cycle 1 requirements)###########################################
        
        ###############################Compressor costs ### axial/centrifugal
        if Pres>=Pin:
            K=PropsSI("ISENTROPIC_EXPANSION_COEFFICIENT", "P", Pin*10**5, "T", Tin, "Hydrogen")
            P2nd = Pin*(Pres/Pin)**(1/3)
            P3rd = Pin*(Pres/Pin)**(1/3)*(Pres/Pin)**(1/3)
            work_c_comp_1 = K/(K-1)*R*Tin/MW_H2*((P2nd/Pin)**((K-1)/K)-1)
            work_c_comp_2 = K/(K-1)*R*Tin/MW_H2*((P3rd/P2nd)**((K-1)/K)-1)
            work_c_comp_3 = K/(K-1)*R*Tin/MW_H2*((Pres/P3rd)**((K-1)/K)-1)
            Work_c_comp = work_c_comp_1+work_c_comp_2+work_c_comp_3
            # Work_c_comp=K/(K-1)*R*Tin/MW_H2*((Pres/Pin)**((K-1)/K)-1) #mechanical energy required for compressor in J/kg (single stage)
            Power_c_comp_1_2=Work_c_comp/1000*m_c_flow_rate_1_2 #mechanical power of the pump in kW
            Power_c_comp_4_2=Work_c_comp/1000*m_c_flow_rate_4_2
            A_c_comp_1_2 = Power_c_comp_1_2/Efficiency_comp  #total power in kW
            A_c_comp_4_2 = Power_c_comp_4_2/Efficiency_comp #total power in kW
            if A_c_comp_1_2>=A_c_comp_4_2:
                A_c_comp=A_c_comp_1_2
            else:
                A_c_comp=A_c_comp_4_2
                
            # print ('work of compressor is', Work_c_comp,'J/kg')
            # print ('Adjusted storage capacity is', H2_c_Cap_Storage, 'kg')
            # print ('flow rate is', m_c_flow_rate_1_2, 'and', m_c_flow_rate_4_2, 'kg/s')
            # print('Total fluid power of compressor', A_c_comp, 'kW')
            Number_c_Compressors=np.floor(A_c_comp/3000) #Number of compressors excluding the last one
            A_c_comp_1 = A_c_comp%3000  #power of the last compressor
            # print('Number of compressors', Number_c_Compressors+1)
            k1=2.2897
            k2=1.3604
            k3=-0.1027
            Compr_c_Cap_Cost=(10**(k1+k2*np.log10(3000)+k3*(np.log10(3000))**2))*Number_c_Compressors
            Compr_c_Cap_Cost_1=(10**(k1+k2*np.log10(A_c_comp_1)+k3*(np.log10(A_c_comp_1))**2))
            Compr_c_Energy_Costs_1=Work_c_comp*H2_c_Cap_Storage*2.8e-7*Energy_cost  #compressor electricity cost in cycle 1
            Compr_c_Energy_Costs_2=Work_c_comp*H2_c_Cap_Storage*Release_efficiency*2.8e-7*Energy_cost #compressor electricity cost assuming in regular charging cycle 
            Total_c_Compr_Cap_Cost = Compr_c_Cap_Cost + Compr_c_Cap_Cost_1
            Total_c_Compr_Cap_Cost = Total_c_Compr_Cap_Cost*(CEPCI_current/CEPCI2001)  ##Inflation
        else:
            Power_c_comp_1_2=0 #mechanical power of the pump in kW
            Power_c_comp_4_2=0
            A_c_comp_1_2 = 0  #total power in kW
            A_c_comp_4_2 = 0 #total power in kW
            Work_c_comp = 0
            Compr_c_Cap_Cost = 0
            Compr_c_Energy_Costs_1 = 0
            Compr_c_Energy_Costs_2 = 0
            Total_c_Compr_Cap_Cost = 0
        # print ('Compressor energy cost is $', Compr_c_Energy_Costs)
        # print ('refrigeration capcost for compressor is $')
        # print('compressor capcost is $', Total_c_Compr_Cap_Cost)
        # print("----------")
        
        ########################################Costs associated with storage tanks
        
        # print("Number of tanks is: ", number_c_of_tanks)
        Storage_c_Tank_Cap_Costs=Cost_c_tank*number_c_of_tanks*Tank_manufacturing
        Storage_c_Tank_Cap_Costs = Storage_c_Tank_Cap_Costs*(CEPCI_current/CEPCI2007)  ##Inflation
        # print('Capcost for storage tank is', Storage_c_Tank_Cap_Costs)
        # print("----------")
        
        ###############################Refrigeration costs estimation adsorption process     
        # print ('pre-conditioning time is', round (t_precondition_hr), 'hr')
        H2_Cap=(capacity)    
        Ref_c_P_net_1_2=-(deltaP_c_net_1_2-Power_c_comp_1_2)  #Refrigeration power in kW from state 1 to state 2 (precondition)
        Ref_c_P_net_4_2=-(deltaP_c_net_4_2-Power_c_comp_4_2)  #Refrigeration power in kW from state 1 to state 2 (normal charging)
        if Ref_c_P_net_1_2 >= Ref_c_P_net_4_2:    
            Net_c_Cooling_Power_Adsorption=Ref_c_P_net_1_2 #Net refrigeration power in kW
        else:
            Net_c_Cooling_Power_Adsorption=Ref_c_P_net_4_2
        # print ('Net Cooling power for refrigeration sizing is',Net_c_Cooling_Power_Adsorption,'kW') #Cooling power in kW
        
        if Net_c_Cooling_Power_Adsorption < 1000:
            A1=-3.53E-09
            A2=-9.94E-06
            A3=3.30E-03
            nc=(A1*(Temp_c**3))+(A2*(Temp_c**2))+A3*Temp_c #Carnot efficiency factor
            COP=(Temp_c/(318-Temp_c))*nc #Coefficient of performance
            B1=24000
            B2=3500
            B3=0.9
            Total_c_Refrig_Cap_Costs_adsorption=(B1+(B2*(Net_c_Cooling_Power_Adsorption/COP)**B3))*(CEPCI_current/550.8)   
        else:
            Total_c_Refrig_Cap_Costs_adsorption=2*10**11*Temp_c**-2.077*(Net_c_Cooling_Power_Adsorption/1000)**0.6
            Total_c_Refrig_Cap_Costs_adsorption = Total_c_Refrig_Cap_Costs_adsorption*(CEPCI_current/CEPCI2017) 
           
        ####Utility for refrigeration
        Utility_c_ref = 4.07*10**7*Temp_c**(-2.669)  #Utility in $/GJ, here, the utility is mostly for energy
        Utility_c_refrigeration_1 = (CEPCI_current/CEPCI2017)*Utility_c_ref*-(deltaE_c_net_1_2-Work_c_comp*H2_c_Cap_Storage/1000)/1e6  
        # print ('refrigerator capital cost for adsorption is $', Total_c_Refrig_Cap_Costs_adsorption)    
        # print("------------")
        
        Utility_c_refrigeration_2 = (CEPCI_current/CEPCI2017)*Utility_c_ref*-(deltaE_c_net_4_2-Work_c_comp*H2_c_Cap_Storage*Release_efficiency/1000)/1e6  
        
            
        ###############################Heating costs desorption process   
        k1=6.9617
        k2=-1.48
        k3=0.3161
        Net_c_Heating_Power_Desorption=detlaP_c_net_2_3/Efficiency_heater ## steam boiler power at 0.7 efficiency in kW
        Number_c_Heaters=np.floor(Net_c_Heating_Power_Desorption/9400) #Number of compressors excluding the last one
        Heater_c_Power_1 = Net_c_Heating_Power_Desorption%9400  #power of the last compressor
        # print('Number of heaters', Number_c_Heaters+1)
        Heater_c_Cap_Cost=(10**(k1+k2*np.log10(9400)+k3*(np.log10(9400))**2))*Number_c_Heaters
        if Heater_c_Power_1 < 1000:
            Heater_c_Cap_Cost_1=(10**(k1+k2*np.log10(1000)+k3*(np.log10(1000))**2))*(Heater_c_Power_1/1000)
        else:
            Heater_c_Cap_Cost_1=(10**(k1+k2*np.log10(Heater_c_Power_1)+k3*(np.log10(Heater_c_Power_1))**2))
        Total_c_Heater_Cap_Cost = Heater_c_Cap_Cost + Heater_c_Cap_Cost_1
        Total_c_Heater_Cap_Cost = Total_c_Heater_Cap_Cost *(CEPCI_current/CEPCI2001)  ##Inflation
        
        Utility_c_Heater=13.28*deltaE_c_net_2_3/1e6 #$13.28/GJ for low pressure steam
        Total_c_Heating_Energy_Costs=Net_c_Heating_Power_Desorption*t_discharge_hr*Energy_cost
        
        # print('heater capcost is $', Total_c_Heater_Cap_Cost)
        
        
        
        
        
        ########################################Operational costs (sized based on cycle 1 requirements)###########################################
        Op_c_Costs_1=Compr_c_Energy_Costs_1 + Utility_c_refrigeration_1+Utility_c_Heater+Total_c_Heating_Energy_Costs
        Op_c_Costs_2=Compr_c_Energy_Costs_2 + Utility_c_refrigeration_2+Utility_c_Heater+Total_c_Heating_Energy_Costs
        Total_c_Cap_Costs = Storage_c_Tank_Cap_Costs + Total_c_Refrig_Cap_Costs_adsorption +Total_c_Compr_Cap_Cost+Total_c_Heater_Cap_Cost
        
        Op_c_Costs = (Op_c_Costs_1 + Op_c_Costs_2 * (cycle_number-1)+maintanance*Total_c_Cap_Costs+wage*360*2)/cycle_number/capacity
        ######################writing costs#####################################################
    
    
        cost_kg [i] = (Total_c_Cap_Costs/capacity + Site_preparation)*Markup
        cost_kg_tank [i] = Storage_c_Tank_Cap_Costs/capacity
        cost_kg_comp [i] = Total_c_Compr_Cap_Cost/capacity
        cost_kg_ref [i] = Total_c_Refrig_Cap_Costs_adsorption/capacity    
        cost_kg_heat [i] = Total_c_Heater_Cap_Cost/capacity
        Op_c_Costs_kg [i] = Op_c_Costs/capacity
        
   
    
   #######################################################################
    ###################Fitting capital####################################################        
    def exp (x,a,b):
        return a*x**b
    
    popt, pcov = curve_fit(exp, capacity_1, cost_kg, maxfev=100000)
    
    #######################################################################
    #######################################################################
    #####################These are the output correlation parameters#########################
    global a_fit
    global b_fit
    
    a_fit=popt[0]
    b_fit=popt[1]
    
    print ('a is', a_fit)
    print ('b is', b_fit)
    print ('***********')
    
    #######################Plotting capital######################
    global fitted_kg
    fitted_kg = exp(capacity_1,a_fit,b_fit)
    plt.figure (1)
    plt.scatter(capacity_1,cost_kg, color='r', label = 'Capex (purchase)')
    plt.plot(capacity_1, fitted_kg, label = 'fitted')
    # plt.plot(capacity_1,cost_kg_tank, color='b', label = 'tank')
    # plt.plot(capacity_1,cost_kg_comp, color='c', label = 'compressor')
    # plt.plot(capacity_1,cost_kg_ref, color='m', label = 'refrigeration')
    # plt.plot(capacity_1,cost_kg_heat, color='y', label = 'heater')   
    
    ####Costmetics######
    a_disp=np.round(a_fit, 2)
    b_disp=np.round(b_fit, 2)
    # plt.ylim(0,np.amax(cost_kg)*2)
    equation = 'y='+str(a_disp)+'x'+'^'+str(b_disp)
    plt.annotate(equation, xy=(np.amax(capacity_1)*0.3 , np.amax(cost_kg)*0.8),size=13) 
    
    plt.xlabel("capacity (kg)")    
    plt.ylabel("Capital cost ($/kg)")
    plt.legend(loc='best', bbox_to_anchor=(1.6, 0.5))
    # plt.legend(loc='best')
    plt.title('Capital')

   
    var_op= [0.01 ,0.5, 5]   #Initial guesses for the parameters, can be flexible
    
    def residual_op(var_op, capacity_1, Op_c_Costs_kg):
    
        a_op=var_op[0]
        b_op=var_op[1]
        c_op=var_op[2]
            
        fit_op_kg = np.exp(a_op*(np.log(capacity_1))**2-b_op*np.log(capacity_1)+c_op)
        return (fit_op_kg - Op_c_Costs_kg)
    
    varfinal_op_fitted,success = leastsq(residual_op, var_op, args=(capacity_1, Op_c_Costs_kg), maxfev=100000)
       
    global a_op_fit
    global b_op_fit
    global c_op_fit  
    a_op_fit=varfinal_op_fitted[0]
    b_op_fit=varfinal_op_fitted[1]
    c_op_fit=varfinal_op_fitted[2] 
   
    fitted_op_kg = np.exp(a_op_fit*(np.log(capacity_1))**2-b_op_fit*np.log(capacity_1)+c_op_fit)
    
    
    a_op_fit_disp=np.round(a_op_fit, 2)
    b_op_fit_disp=np.round(b_op_fit, 2)
    c_op_fit_disp=np.round(c_op_fit, 2) 
    
    equation_op = 'y='+'exp('+str(a_op_fit_disp)+'(ln(x))^2-'+str(b_op_fit_disp)+'ln(x)+' + str(c_op_fit_disp) +')'
    #######################################################################
    #######################################################################
    
    #####################These are the output correlation parameters for operational#########################
    print ('a_op is', a_op_fit)
    print ('b_op is', b_op_fit)
    print ('c_op is', c_op_fit)
    print ('***********')
    
    
    #######################Plotting operational######################  
    plt.figure (2)
    plt.plot(capacity_1, fitted_op_kg, label = 'fitted')
    plt.scatter(capacity_1,Op_c_Costs_kg, color='r', label = 'Annual opex')    
    plt.xlabel("capacity (kg)")    
    plt.ylabel("Operational cost ($/kg)")
    plt.annotate(equation_op, xy=(np.amax(capacity_1)*0.3 , np.amax(Op_c_Costs_kg)*0.8),size=13)     
    plt.legend(loc='best', bbox_to_anchor=(1.6, 0.5))
    # plt.legend(loc='best')
    plt.title('Annual operational')
    plt.show()